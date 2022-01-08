import requests
from bs4 import BeautifulSoup

from selenium import webdriver
import PyPDF2
import os
import openpyxl

def test01():
    url = 'https://stackoverflow.com/questions'
    repose = requests.get(url)

    soup = BeautifulSoup(repose.text, 'html.parser')
    questions = soup.select('.question-summary')
    for question in questions :
        print(question.select_one('.question-hyperlink').getText())
        print(question.select_one('.vote-count-post').getText())

# browser = webdriver.Chrome()
def test02():
    # step1 : open webpage
    url = 'https://github.com'
    global browser
    browser.get(url)
    sign_up = browser.find_element_by_link_text('Sign up')
    sign_up.click()
    sign_in = browser.find_element_by_link_text('Sign in →')
    sign_in.click()

    # step2 : enter message
    username_box = browser.find_element_by_id('login_field')
    username_box.send_keys('username')
    password_box = browser.find_element_by_id('password')
    password_box.send_keys('password')
    password_box.submit()

    # step3 : quit webpage
    assert 'gao-haoyu' in browser.page_source
    browser.quit()

def test03():
    url = 'https://baidu.com'
    browser.get(url)
    sign = browser.find_element_by_link_text('登录')
    sign.click()

def test04():

    with open('first.pdf', 'wb') as file1, open('second.pdf', 'wb') as file2:
        with open('./C语言C++常见面试题（含答案）.pdf', 'rb') as source:
            reader = PyPDF2.PdfFileReader(source)
            page = reader.getPage(0)
            writer = PyPDF2.PdfFileWriter()
            writer.addPage(page)
            writer.write(file1)
            writer.write(file2)


def test06():
    wb = openpyxl.load_workbook('111.xlsx')
    print(wb.sheetnames)
    sheet1 = wb['Sheet1']
    print(sheet1.max_row)
    print(sheet1.max_column)
    sheet1.append([10, 11, 12])
    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            print(sheet1.cell(row, col).value)
    wb.save('222.xlsx')

